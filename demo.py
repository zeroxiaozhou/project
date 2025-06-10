import requests
import time
from datetime import datetime, timedelta
from sqlalchemy import create_engine, text, inspect
# import datetime
import pandas as pd
from tqdm import trange
import os
import emoji
import openpyxl
from openpyxl.styles import Alignment, Font
from win32com.client import Dispatch
from copy import copy
# import shutil
# import requests
# import json
# import docx
import random
# import base64
import hashlib
# from docx import Document
# from Crypto.Cipher import DES
# from Crypto.Util.Padding import pad
# from Crypto.Util.strxor import strxor
import execjs
# import time
# import rsa
# import Crypto.PublicKey.RSA
# import Crypto.Random
# from Crypto.Cipher import PKCS1_v1_5
import browser_cookie3
# import pprint
# import re
# import shutil
# import js2py
# import execjs
import random
# import ddddocr
# import string
# import os
# import datetime
import pymssql
# from lxml import etree
# import zipfile
# import concurrent.futures
# from concurrent.futures import ThreadPoolExecutor, as_completed
import traceback
# import urllib
# from reportlab.lib.pagesizes import letter
# from reportlab.pdfgen import canvas
# from PIL import Image
# 图片转pdf
# def imgtopdf(input_paths, outputpath):
#     maxw, maxh = Image.open(input_paths).size
#     pdf_w, pdf_h = (480.28, 702.78)
#     c = canvas.Canvas(outputpath, pagesize=(pdf_w, pdf_h))
#     if maxw / pdf_w > maxh / pdf_h:
#         c.drawImage(input_paths, 0, (pdf_h - maxh * pdf_w / maxw) / 2, pdf_w, maxh * pdf_w / maxw)
#     else:
#         c.drawImage(input_paths, (pdf_w - maxw * pdf_h / maxh) / 2, 0, maxw * pdf_h / maxh, pdf_h)
#     c.showPage()
#     c.save()
# imgtopdf(r"C:\Users\XYCX\Desktop\营业执照v.jpg", r"C:\Users\XYCX\Desktop\营业执照v.pdf")
# engine = create_engine('mssql+pymssql://sa:Xy202204@LAPTOP-HHMD1A48\\XYCX/证件合规信息查询?charset=utf8', echo=False)
# engine = create_engine('mysql+pymysql://root:Xy=202212@47.93.173.200:3306/bak_yueyuechuxing?charset=utf8', echo=False)
# df = pd.read_sql('SELECT * FROM `Account`', con=engine)
# print(df)
# engine = create_engine('mssql+pymssql://sa:sjz123@SHUJUZHU\MSSQLSERVER2/订单热点1?charset=utf8', echo=False)

# 示例用法
# image_to_pdf(r"C:\Users\XYCX\Desktop\营业执照v.jpg", r"C:\Users\XYCX\Desktop\营业执照v.pdf")

# url = 'https://nc.taxiedu.com:82/web/PxRyDetail.aspx?Xm=黎达才&sfzh=362525199806161516'
# header = {
#     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 Edg/110.0.1587.50'}
# res = requests.session().get(url, headers=header)
# print(res.text)
# engine = create_engine('mssql+pymssql://sa:Xy202204@LAPTOP-HHMD1A48\XYCX/python?charset=utf8',echo=False)
# file = os.listdir(r'D:\呼我')
# for i in file:
#     datn = pd.read_excel(rf'D:\呼我\{i}')
#     print(f'文件读取完成(u‿ฺu✿ฺ)')
#     datn.columns = ['司机id', '司机姓名', '司机手机号', '司机类型', '业务类型', '激活时间', '状态', '星级评价',
#                     '品牌', '注册时间', '审核通过时间',
#                     '首次出车时间', '最后一次出车时间', '首次完单时间', '最后一次完单时间', '所属城市', '运力公司',
#                     '出车时长', '服务时长', '有效出车时长',
#                     '高峰期出车时长', '高峰期有效时长', '出勤达标天数', '应答量(总)', '应答量(快车)', '应答量(专车)',
#                     '完单量(总)', '完单量(快车)',
#                     '完单量(专车)', '完成支付订单量', '出车天数', '完单天数', '考核完单率', '订单应付金额',
#                     '订单总金额', '司机取消订单量', '系统关闭订单数',
#                     '投诉订单量', '仅听预约单时长(小时)', '听顺路单时长(小时)', '行程费退款金额', '远程调度费退款金额',
#                     '附加费退款金额', '行程费',
#                     '远程调度费', '附加费', '司机抽佣行程费用', '司机订单流水', '首单奖', '冲单奖', '所有活动奖励',
#                     '早高峰接单量', '早高峰完单量',
#                     '晚高峰接单数', '晚高峰完单数', '服务分', '车队', '车牌号', '平台活动奖励', '近30天差评率',
#                     '近30天客诉率', '成交率']
#     datn['数据日期'] = '2023-04-22'
#     datn['司机id'] = datn['司机id'].apply(lambda x: str(x).replace('.0', ''))
#     datn.to_sql('司机考核数据日数据', con=engine, if_exists='append', index=False)
#     print(f'写入数据库完成(✿◡‿◡)')

# a = '%25E6%259E%2597%25E5%2590%259B%25E5%2586%259B'
# b = urllib.parse.unquote(a)
# print(urllib.parse.unquote(b))


# print(str,sign)
# url = 'http://zzjk.zztaxi.cn:8090/netCarMonitor/interface/auth/serverVerify.do'


# engine = create_engine('mssql+pymssql://sa:Xy202204@LAPTOP-HHMD1A48\XYCX/证件合规信息查询?charset=utf8',echo = False)
# data = pd.read_excel(r'D:\双证合规\佛山市\佛山市人证推送04-04.xlsx')
# new_data = data[['司机ID','车牌号','城市','品牌','身份证号','司机姓名','车辆运输证号','证件状态','更新日期']]
# # print(data)
# print(new_data)

# 郑州接口文档
# import requests
# url = 'http://zzjk.zztaxi.cn:8090/netCarMonitor/interfaceQuery/getVerificationStatus.do'
# header = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36",
#     "interfaceToken": "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiJ9.eyJzdWIiOiJ7XCJpbnRlcmZhY2VVc2VyTmFtZVwiOlwi5rGf5Y2X5Ye66KGMXCIsXCJ1c2VyTm9cIjpcImE3Yjk4MmM1YTZhMjQ3MjdhZmU5M2M5Y2ViOTIxNTFjXCIsXCJjb21wYW55TmFtZVwiOlwi5rGf5Y2X5Ye66KGMXCJ9IiwiaWF0IjoxNjg1OTQwMjE2LCJleHAiOjE1NTY4ODU5NDAyMTZ9.oVIzpjsgAkJJjLgRPn4cqqUiAMMHk5c2QtG_Vdn_t_fiY3yZ4r69BbzMg712eOjeuyCJvHYYpWpFOMZfknFaoA"}
# param = {"key": "MIGfMA0GCSqGSIb3DQEBAQUAA4GNADCBiQKBgQDfqni8K9IAUXXIHL6fHoBMrgxIYNU2YgXOJa42jXlJp6Roj2kCImYm/5utcibxFlUsekhZUSIxljY71Jn6LSEDKn2wClXO2KcdmXJy96cbyeR0YI/mJ6a/08OCsPhJ7g+t0Mi+ut7GrBmNSt3FOrFDLrIGqnjC0hwHDbQRnXYyBQIDAQAB",
#          "vehicleNo": "豫ADG6167", "idCardNo": "410222198208024012"}
# res = requests.get(url, headers=header, params=param)
# # status: 1是合规，0是不合规
# print(res.text)
# import random
# import time
# import hashlib
# import requests
# import base64
# from Crypto.Signature import PKCS1_v1_5
# from Crypto.Hash import SHA256
# from Crypto.PublicKey import RSA
#
# ALGORITHM_NAME = "RSA"
# MD5_RSA = "MD5withRSA"
# def get_token():
#     apiId = "rf9c552da4b6dcd814"
#     apiSecret = "9b9782c2cbc74d9d865727db376ccfcf"
#     privateKey = "MIICdwIBADANBgkqhkiG9w0BAQEFAASCAmEwggJdAgEAAoGBAN+qeLwr0gBRdcgcvp8egEyuDEhg1TZiBc4lrjaNeUmnpGiPaQIiZib/m61yJvEWVSx6SFlRIjGWNjvUmfotIQMqfbAKVc7Ypx2ZcnL3pxvJ5HRgj+Ynpr/Tw4Kw+EnuD63QyL663sasGY1K3cU6sUMusgaqeMLSHAcNtBGddjIFAgMBAAECgYAUhzHjm2X/z3ou7qx0MDl4UDUiY3jOL/r2a7Dsotlx8Cf/zMHHh162z5j7N1HpqLISjfqb7/1ibbX2kdG8C25PDxyBmEf+Wa9Cgra0JTPq9gnQYRvkMBCzyTqPH+OAnWSyV3bm+wwE0rWNOFm73rmeYLWOz0eKVTFF3G15KHwBNQJBAPoxIxybAi3Vy6fP+Nknn3bmLE1Jtp8dwt9m6a6eW4seLtceHkt6gGo+7TMxgz4Nu/43AcTPXkx2iFpe5A527+8CQQDk27F0UqCsc5gVlsdpnTpxWSwotog7FIHJvoOKLU599LMFBQxewCrnDpguMHDIm/2T/TtA5WBns2jsIu6SOYlLAkEA5p6ImQOhXJKoKUWhQrotWbINwChkeAM88CSy3s0F4RSvZIdUsYp3+HeMuhW3vml2knwt2zay25SfV34EhfjIbQJARrXegVtaS443qkv49xfeS9FKhJXJR7/RTh0wFUxkWSR2/5EMvmXPm651tKfA4SrZUZVHboiwnbngLD2qysE+OwJBAPC+g7VBr0FWpU0tbkePXhlYzlHnftEvIyYWYndJd6uH+Su2GYxG99MCGJAMCpR00v8A7vU7E4iZgX1egI8ZEIY="
#     nonce = ''.join(random.sample('zyxwvutsrqponmlkjihgfedcba0123456789', 10))
#     req_time = str(round(time.time()*1000))
#     str_to_sign = apiId + '&' + req_time + '&' + nonce
#     sign = get_SHA256_str(str_to_sign)
#     map = {
#         'apiId': apiId,
#         'apiSecret': apiSecret,
#         'nonce': nonce,
#         'sign': sign,
#         'time': req_time
#     }
#     sb = ''
#     for key in map.keys():
#         val = map[key]
#         if val.strip() != '':
#             sb += key + '=' + val.strip() + '&'
#     # 示例用法
#     api_sign = sign_str(sb, privateKey)
#     map['apiSign'] = api_sign
#     url = 'http://zzjk.zztaxi.cn:8090/netCarMonitor/interface/auth/serverVerify.do'
#     print("第一个map===>>", map)
#     response = requests.post(url, json=map)
#     print(response.json())
#     map['code'] = response.json()['data']
#     # print("map===>>", map)
#     # url_token = 'http://zzjk.zztaxi.cn:8090/netCarMonitor/interface/auth/getToken.do'
#     # token_response = requests.post(url_token, data=json.dumps(map), headers=headers)
#     # token_body = token_response.json()
#     # print("tokenBody===>>"+str(token_body))
#     return map
# def sign_str(data_str, private_key_str):
#     # 将私钥转换成 RSA 密钥对象
#     private_key_bytes = base64.b64decode(private_key_str)
#     pkey = RSA.import_key(private_key_bytes)
#     # 使用 PKCS1_v1_5 签名算法进行签名
#     h = SHA256.new(data_str.encode('utf-8'))
#     signer = PKCS1_v1_5.new(pkey)
#     signature_bytes = signer.sign(h)
#     # 返回 Base64 编码的签名字符串
#     signature_str = base64.b64encode(signature_bytes).decode('utf-8')
#     return signature_str
#
# def get_SHA256_str(str):
#     messageDigest = hashlib.sha256()
#     messageDigest.update(str.encode('utf-8'))
#     hash = messageDigest.digest()
#     return hash.hex()
# get_token()


# import pprint
# import pandas as pd
# import requests
# import urllib
# import time
# name = urllib.parse.quote('鲁B')
# print(name)
# quit()
# url = 'http://27.223.104.252:18080/wycht/vehicle/rhBaseVehicle/list?pageNum=1&pageSize=15&params%5BvehicleNo%5D=%7B%22condition%22%3A%22LIKE%22%2C%22value%22%3A%22%E9%B2%81BD68599%22%7D'
# headers = {
#     "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36 Edg/114.0.1823.41",
#     "Authorization": f"Bearer "
# }
# res = requests.get(url, headers=headers)
# print(res.status_code)
# pprint.pprint(res.json())
# today = str(time.strftime('%Y-%m-%d', time.localtime()))
# with open(r'C:\Users\XYCX\Desktop\青岛所有司机.txt', 'r', encoding='utf16') as f:
#     data = json.loads(f.read())
# df = pd.DataFrame()
# num = 0
# for i in data['rows']:
#     df = df.append(pd.DataFrame(i, index=[0]))
#     num += 1
#     print(num)
# df.to_excel(r'D:\双证合规\青岛市\青岛市人证%s.xlsx' % today[5::], sheet_name='人证', index=False)

# txt存进excel
# with open(r'C:\Users\XYCX\Desktop\青岛网约车证.txt', 'r', encoding='utf16') as f:
#     data = json.loads(f.read())
# num = 0
# df = pd.DataFrame()
# for i in data['rows']:
#     df = df.append(pd.DataFrame(i, index=[0]))
#     num += 1
#     print(num)
# df.to_excel(r'D:\双证合规\青岛市\青岛市车证%s.xlsx' % today[5::], sheet_name='车证', index=False)


# df = pd.read_excel(r'C:\Users\XYCX\Desktop\广州市注册报备.xlsx')
# file = docx.Document(r'C:\Users\XYCX\Desktop\喜行约车协议.docx')
# for i in range(len(df)):
#     name = df['姓名'][i]
#     id = df['身份证号'][i]
#     driver_id = df['司机ID'][i]
#     # print(name, id)
#     file.paragraphs[2].add_run(name)
#     file.paragraphs[3].add_run(id)
#     file.save(rf'D:\喜行约车司机协议\{driver_id}.docx')
#     file = docx.Document(r'C:\Users\XYCX\Desktop\喜行约车协议.docx')


# df = pandas.read_excel(r'C:\Users\XYCX\Desktop\喜行广州推送司机.xlsx')
# file_all = os.listdir('D:/司机协议PDF/')
# for i in df.index:
#     try:
#         old_name = df['司机姓名'][i]+df['身份证号'][i]
#         new_name = df['司机ID'][i]
#         print(old_name, new_name)
#         os.rename(rf'D:/new_司机协议PDF/{old_name}.pdf', rf'D:/司机协议PDF/{new_name}.pdf')
#     except Exception:
#         continue
#
# # 设置每个文件夹包含的PDF文件数
# batch_size = 100
#
# # 遍历当前目录下的所有PDF文件
# pdf_files = [f for f in os.listdir('D:/司机协议PDF')]
# print(pdf_files)
# # 根据batch_size新建文件夹，并移动PDF文件到对应文件夹
# for i, pdf_file in enumerate(pdf_files):
#     # 计算文件夹序号
#     folder_num = i // batch_size + 1
#     # 构造文件夹名
#     folder_name = fr'D:\司机协议分批\司机协议{folder_num:03d}'
#     # 新建文件夹（如果不存在）
#     os.makedirs(folder_name, exist_ok=True)
#     # 移动PDF文件到对应文件夹
#     shutil.move(rf'D:\司机协议PDF\{pdf_file}', os.path.join(folder_name, pdf_file))

# df = pd.read_excel(r'C:\Users\XYCX\Desktop\喜行未注册成功名单.xlsx')
# file_all = os.listdir('D:/司机协议PDF/')
# for i in df.index:
#     new_name = df['司机ID'][i]
#     print(new_name)
#     shutil.move(rf'D:\星徽司机协议\{new_name}.pdf', rf'D:\星徽司机协议PDF\{new_name}.pdf')


# import requests
# import time
# import pandas as pd
# from sqlalchemy import create_engine
#
# def result(x):
#     band = str(x)
#     if str(x) == "2":
#         band = "帮邦行"
#     return band
# df = pd.DataFrame()
# url = 'https://mozhu.amap.com/api/getCityList?cpCode=9422'
# headers = {
#     'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36',
#     'cookie': '_uab_collina=168128461039078602647667; cna=KEy9HDjIkVICAXQWAf8TMe21; passport_login=NDMyMzAxMjM4LGFtYXBfMTU5OTk5MDkxMTZCZ1lXcWp4TkosNWR4ZmhncmJ6ZXdjYnRjanFuaWZ1anMya2pzdW5sN2osMTY4NjgwOTIzNyxZelk1TUdVeU5qZGlPVFZqT0dZMU1tTXdOVEUyWkRWaVpHSXdZVFpoTWprPQ%3D%3D; xlly_s=1; JSESSIONID=19e8d57f94a8407659fff4ca3e5b07ba2bf13b606989b5d5354ce61fd3f03bc276ae6eced169746bd969fd22f030b0499add3964a6ccf0af309398dc82a08cad0fc4b7822f79bda81e0395a092fff65b6c0a2195d72278536e4aa787d6574ab1734d1ec48b44a9a7a744b354639663cec032b6c4ab764bc8e814e526667dbf44eb6a572e9379f8da871328a47e43e8d416ab4e812445c33cf9eb96ad87de832bc46c10b775913bb0c62f1f79b5d151595c970f510772ec8ce95b9e6e93a8b4a11e99c1ac8ad5f1f75d1ad2fc6489db2d873daa2a09a8a95c91f14cf7c878cdfc7f0f132548caf7fe2a557f950bc45aa1e27c660b2aaaf7ff2bc3c6d3f515b9692d7731184c2354876c54cc9c9234a0e970b21dafdf635b4a475fdf36bf67f2f678afa65931016b8780f37b5b7d674aae462becc80ab09127eee77ac46b555df4c7b6b57c0fac9ace6e22840a6ab93ef1ab08e525e0aacf530da97645f9e6ab3ffabf513eb1d50da555544ac0114c55d350fa5d2f6cf815d3ef2eab11d090570467a633a21e0bc89624b5b0a1a922915aa6f7e897b3d6873636a4179d78ccfd9dfc8686ef3d7b2a08320826273915c4c5d34920edcd3d59dde3b9d84a988b5ccd971bc38ca0fd6facd6d74cda985f99085b0e5918abf1ee98e7896dac909a6417ca2f2781773013d8cf4daa82d9b3c80f312715c26add4b480ecd55b1654dffb678cf7955621390093a128b71fa0017eb0a2f246f0aadfe7952889c6b5b3790d41dbf01255a60593167c8a1a087c3b4ef2e529633ffeb07ffa0352ceb0e74e67c1fd2f39fa5c0661fb0b9008fe0d610320d5554b78472a2432fe9e9400719a950c35bc37a6272362d1f50f2a07201f3cbfcd596a88da54213944959438d5fff2fe37a39c8418c0492b8edc3ce36b5455573556ab89bf8c2bb85659a78f39f1ee8a5ecfd220d4225f1bd396162e7f59d56; tfstk=cUrABZTiO_fcLqfi4mQkQpop_V2AZ5pxsKG9XoIKUKqcbb-Oi1E3vv-FlAlEMTC..; l=fBrahikeNM3Xx7MaBOfwPurza77OSIRAguPzaNbMi9fP_JfJ5IcfW11BpD8vC3GVFsgWR3lDK4dwBeYBqIcIdJEZpW9Rw_HmnmOk-Wf..; isg=BHp6mJlvynjxvEav6sQXmn_cy6CcK_4FDnNRAIR_wI1BdwexSLsCFHmRxQOrZ3ad'
# }
# res = requests.get(url, headers=headers)
# print(res.json()["data"])
# df = df.append(pd.DataFrame(res.json()["data"]), ignore_index=True)
# df['cityStatus'] = df['cityStatus'].apply(result)
# df.rename(columns={"cityStatus": "品牌", 'label': '所属城市', "value": "行政区划代码"}, inplace=True)
# engine = create_engine('mssql+pymssql://sa:Xy202204@LAPTOP-HHMD1A48\XYCX/python?charset=utf8', echo=False)
# df.to_sql('品牌完单城市', con=engine, if_exists='append', index=False)

# engine = create_engine('mssql+pymssql://sa:Xy202204@LAPTOP-HHMD1A48\XYCX/python?charset=utf8', echo=False)
# datn = pd.read_excel(r"C:\Users\XYCX\Downloads\司机考核列表导出.xlsx")
# print(f'文件读取完成(u‿ฺu✿ฺ)')
# datn.columns = ['司机id', '司机姓名', '司机手机号', '司机类型', '业务类型', '激活时间', '状态', '星级评价',
#                 '品牌', '注册时间', '审核通过时间',
#                 '首次出车时间', '最后一次出车时间', '首次完单时间', '最后一次完单时间', '所属城市', '运力公司',
#                 '出车时长', '服务时长', '有效出车时长',
#                 '高峰期出车时长', '高峰期有效时长', '出勤达标天数', '应答量(总)', '应答量(快车)', '应答量(专车)',
#                 '完单量(总)', '完单量(快车)',
#                 '完单量(专车)', '完成支付订单量', '出车天数', '完单天数', '考核完单率', '订单应付金额',
#                 '订单总金额', '司机取消订单量', '系统关闭订单数',
#                 '投诉订单量', '仅听预约单时长(小时)', '听顺路单时长(小时)', '行程费退款金额', '远程调度费退款金额',
#                 '附加费退款金额', '行程费',
#                 '远程调度费', '附加费', '司机抽佣行程费用', '司机订单流水', '首单奖', '冲单奖', '所有活动奖励',
#                 '早高峰接单量', '早高峰完单量',
#                 '晚高峰接单数', '晚高峰完单数', '服务分', '车队', '车牌号', '平台活动奖励', '近30天差评率',
#                 '近30天客诉率', '成交率']
# datn['数据日期'] = '2024-05-30'
# datn['司机id'] = datn['司机id'].apply(lambda x: str(x).replace('.0', ''))
# datn.to_sql('司机考核数据日数据', con=engine, if_exists='append', index=False)

# da = pd.read_excel(r"C:\Users\XYCX\Desktop\新司机报考.xlsx")
# data = da[['司机ID', '车牌号', '城市', '品牌', '身份证号', '姓名', '车辆运输证号', '人证']]
# engine = create_engine('mssql+pymssql://sa:Xy202204@LAPTOP-HHMD1A48\XYCX/证件合规信息查询?charset=utf8', echo=False)
# da.to_sql('新司机报考', con=engine, if_exists='append', index=False)
#
# df = pd.DataFrame()

# 自动获取本地cookie
# headers = {
#     'cookie': '',
#     '_admin_eid': '', '_admin_session_eid': '', '_admin_sign': '76828b6260c0588b86dff532333faab9',
#     '_admin_tk': '', '_admin_ts': str(round(time.time()))}
# cj = browser_cookie3.edge(domain_name='admin.yueyuechuxing.cn')
# for cookie in cj:
#     if cookie.name == 'stk':
#         headers['_admin_tk'] = cookie.value
#     if cookie.name == '_tenant_id':
#         headers['_admin_eid'] = cookie.value
#         headers['_admin_session_eid'] = cookie.value
#         headers['cookie'] = headers['cookie'] + cookie.name + '=' + cookie.value + ';'
#     else:
#         headers['cookie'] = headers['cookie'] + cookie.name + '=' + cookie.value + ';'
# print(headers)
# url = 'https://admin.yueyuechuxing.cn/admin/v1/sv/driver/findSummaryList'
# data = {'adcode': "610900", 'pageNum': 1, 'pageSize': 10}
# res = requests.post(url, headers=headers, json=data)
# print(res.json())
# 36位时间戳
# def base36encode(number):
#     ALPHABET = '0123456789abcdefghijklmnopqrstuvwxyz'
#     base36 = ''
#     sign = ''
#
#     if number < 0:
#         sign = '-'
#         number = -number
#
#     if 0 <= number < len(ALPHABET):
#         return sign + ALPHABET[number]
#
#     while number != 0:
#         number, i = divmod(number, len(ALPHABET))
#         base36 = ALPHABET[i] + base36
#
#     return sign + base36
#
# timestamp = int(datetime.datetime.now().timestamp() * 1000)
# timestamp_str = str(timestamp)
# timestamp_base36 = base36encode(int(timestamp_str))

# func_js = """
# function add(){
#     var a = 1;
#     var b = 2;
#     return a+b
# }
# """
# add1 = js2py.eval_js(func_js)
# print(add1())

# 全国加密逆向
"""
import string
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad
from Crypto.PublicKey import RSA

# 全国网站逆向
def getRamNumber(digit):
    result = ''
    for _ in range(digit):
        result += random.choice(string.ascii_uppercase + string.digits)
    return result.upper()

AES_KEY_IV  = '1234567812345678'
def aesEncrypt(key, body_text):
    key = key.encode()
    iv = AES_KEY_IV.encode()
    cipher = AES.new(key, AES.MODE_CBC, iv)
    encrypted = cipher.encrypt(pad(body_text.encode(), AES.block_size))
    return encrypted.hex()

def rsa_encrypt(body_text):
    key = RSA.importKey(RSA_PUBLIC_KEY)
    cipher = PKCS1_v1_5.new(key)
    encrypted = cipher.encrypt(body_text.encode('utf-8'))
    return base64.b64encode(encrypted).decode('utf-8')

headers = {
    'Accept': 'image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
    'Cache-Control': 'no-cache',
    'Connection': 'keep-alive',
    'Pragma': 'no-cache',
    'Referer': 'http://ysfw.mot.gov.cn/NetRoadCGSS-web/information/query',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36 Edg/119.0.0.0',
}
random_number = random.random()
sess = requests.session()
# 获取验证码图片
response = sess.get(f'http://ysfw.mot.gov.cn/NetRoadCGSS-web/captcha.jspx?{random_number}', headers=headers).content
# 使用快识别解析验证码
base64_data = base64.b64encode(response)
b64 = base64_data.decode()
data = {"username": 'Xy123', "password": 'Xy123456', "typeid": 3, "image": b64}
result = json.loads(requests.post("http://api.ttshitu.com/predict", json=data).text)
if result['success']:
    captcha = result["data"]["result"]
else:
    captcha = result["message"]
print(captcha)
# 请求数据
data = {
    "staffName": "赵祥鹤",
    "idType": "1",
    "idCard": "340122198511151811",
    "Page": 1,
    "PageSize": 10,
    "captcha": captcha
}
RSA_PUBLIC_KEY="-----BEGIN PUBLIC KEY-----\nMFwwDQYJKoZIhvcNAQEBBQADSwAwSAJBAJKTV32+OIzBMTDQreJGwjDe8/88a6QpeKSWdivyQkvIwq8d0BL1cOiSujKZP+G+3LGBfha2B6O0EiLT1ArKgC0CAwEAAQ==\n-----END PUBLIC KEY-----"
aesKey = getRamNumber(16)
requestData = aesEncrypt(aesKey, json.dumps(data))
encrypted = rsa_encrypt(aesKey)
param = {"requestData": requestData, "encrypted": encrypted}
# 取得加密参数requestData，encrypted
print(param)
# 注释掉的是第二种方法，读取js文件进行解密取得请求参数
# with open(r'./人证车证查合规/全国逆向.js', 'r', encoding='utf-8') as f:
#     Decryption = f.read()
#
# ctx = execjs.compile(Decryption)
# print(ctx.call('postParamMake', data))
# param = ctx.call('postParamMake', data)
res = sess.post('http://ysfw.mot.gov.cn/NetRoadCGSS-web/information/staffinfoquery', headers=headers, json=param)
print(res.json())"""

"""
# 太原人证网站逆向
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36 Edg/119.0.0.0',
}
n = {
    "cardno": "140122198801272817",
    "name": "郑旭明",
    "pagenum": 1
}
e = "LTPUCZC_getQualificationByCardnoAndName"
with open(r'./人证车证查合规/太原逆向.js', 'r', encoding='utf-8') as f:
    Decryption = f.read()
ctx = execjs.compile(Decryption)
print(ctx.call('main', n))
param = ctx.call('main', n)
response = requests.post(
    'https://taxi.jtzyzg.org.cn/CZCJSY/restservices/http/single/query',
    headers=headers,
    json=param
)
print(response.json())
"""
"""
# 江西网站逆向
# 获取当前时间戳
timestamp = str(int(time.time()))
# 要加密的字符串
input_str = 'wtkj2020wtkj2020' + timestamp
# 创建 MD5 对象
md5 = hashlib.md5()
# 更新对象以要加密的数据
md5.update(input_str.encode('utf-8'))
# 获取加密后的结果
token = md5.hexdigest()
print("MD5 加密后的结果:", token)
"""
# headers = {
#     'Accept': '*/*',
#     'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
#     'Connection': 'keep-alive',
#     'Content-Type': 'application/x-www-form-urlencoded',
#     'Cookie': 'connect.sid=s%3ABZrCSY0oT7qUI6HeoHf_Ccs5rAz2QlUp.SzHQgzFo6qa4canoMbL%2FCUjNjMx7uBn%2FvaHRbuF4Jno',
#     'Origin': 'https://www.zhczqc.org.cn:10203',
#     'Referer': 'https://www.zhczqc.org.cn:10203/archives',
#     'Sec-Fetch-Dest': 'empty',
#     'Sec-Fetch-Mode': 'cors',
#     'Sec-Fetch-Site': 'same-origin',
#     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0',
#     'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Microsoft Edge";v="120"',
#     'sec-ch-ua-mobile': '?0',
#     'sec-ch-ua-platform': '"Windows"',
# }
# engine = create_engine('mssql+pymssql://sa:Xy202204@LAPTOP-HHMD1A48\XYCX/证件合规信息查询?charset=utf8', echo=False)
# sql = "SELECT * FROM [证件合规信息查询].[dbo].[每日需要查询数据] WHERE 城市='珠海市'and 品牌='喜行约车'"
# datn = pd.read_sql(sql, con=engine)
# sess = requests.session()
# df = pd.DataFrame()
# df2 = pd.DataFrame()
# for i in trange(892, len(datn)):
#     data = f'jsonData={{"carrerNo": "{str(datn["身份证号"][i])}"}}'
#     # data = 'jsonData={"carrerNo":"432502198508152311"}'
#     response = sess.post(
#         'https://www.zhczqc.org.cn:10203/archives/wyc2/getCheckJsyByCareerNo',
#         headers=headers,
#         data=data,
#     ).json()
#     if response['status'] == 0:
#         # 将json数据转换为DataFrame
#         data = pd.DataFrame.from_dict(response['value'], orient='index').T
#         # 将DataFrame添加到df中
#         df = df.append(data, ignore_index=True)
#         df.to_excel(r'D:\双证合规\珠海市\珠海市人证查询01-15-2.xlsx', index=False)
#     else:
#         print(response)
#         df2 = df2.append(pd.DataFrame({'司机ID': datn['司机ID'][i], '车牌号': datn['车牌号'][i], '城市': datn['城市'][i], '品牌': datn['品牌'][i],
#                               '身份证号': str(datn['身份证号'][i]), '姓名': datn['姓名'][i], '车辆运输证号': datn['车辆运输证号'][i]}, index=[0]))
#         df2.to_excel(r'D:\双证合规\珠海市\珠海市不合规人证01-15.xlsx', index=False)
#     time.sleep(0.4)

# print("\033[31m我是小杨我就这样\033[0m")
# print("\033[32m我是小杨我就这样\033[0m")
# print("\033[33m我是小杨我就这样\033[0m")
# print("\033[34m我是小杨我就这样\033[0m")
# print("\033[35m我是小杨我就这样\033[0m")
# print("\033[36m我是小杨我就这样\033[0m")
# print("\033[37m我是小杨我就这样\033[0m")
# print(f'\033[31m写入数据库完成(✿◡‿◡)\033[0m')
# print(f'\033[36m======================完成======================\033[0m')

# # 读取第一张表
# df1 = pd.read_excel(r"C:\Users\XYCX\Desktop\汇总表.xlsx")
# # 打开Excel文件
# workbook = openpyxl.load_workbook(r"C:\Users\XYCX\Desktop\附件4：网约出租汽车驾驶员服务质量信誉档案.xlsx")
# # 获取第二个工作表
# second_sheet = workbook.worksheets[0]
# second_sheet1 = workbook.worksheets[1]
# second_sheet2 = workbook.worksheets[2]
# second_sheet3 = workbook.worksheets[3]
# second_sheet4 = workbook.worksheets[4]
# for i in trange(len(df1)):
#     driver = str(df1['司机id'][i]).split('.')[0]
#     name = df1['姓名'][i]
#     ID = df1['身份证号'][i]
#     riqi = df1['人证领取日期'][i]
#     # 将name和ID放在元组中
#     name_tuple = (name, name, name, name)
#     # print(driver, name, ID, riqi)
#     # continue
#     # 获取A3单元格的值
#     second_sheet['H2'] = driver
#     second_sheet['H2'].alignment = Alignment(vertical='center', horizontal='center')
#
#     second_sheet['H3'] = '王跃洋'
#     second_sheet['H3'].alignment = Alignment(vertical='center', horizontal='center')
#
#     second_sheet['H4'] = riqi
#     second_sheet['H4'].alignment = Alignment(vertical='center', horizontal='center')
#
#     second_sheet['D11'] = '海口喜行网络科技有限公司'
#     second_sheet['D11'].alignment = Alignment(vertical='center', horizontal='center')
#
#     second_sheet['D12'] = name
#     second_sheet['D12'].alignment = Alignment(vertical='center', horizontal='center')
#     cell_font = Font(underline=None)
#     second_sheet['D12'].font = cell_font
#
#     second_sheet['D14'] = ID
#     second_sheet['D14'].alignment = Alignment(vertical='center', horizontal='center')
#
#     second_sheet1['B5'].value, second_sheet2['B5'].value, second_sheet3['B5'].value, second_sheet4['B5'].value = name_tuple
#     second_sheet1['H5'] = ID
#     second_sheet2['I5'] = ID
#     second_sheet3['I5'] = ID
#     second_sheet4['F5'] = ID
#
#     workbook.save(rf"C:\Users\XYCX\Desktop\档案汇总3\{df1['姓名'][i]}.xlsx")


# file_path = r"C:\Users\XYCX\Desktop\档案汇总3"
# for filename in os.listdir(file_path):
#     file = fr"C:\Users\XYCX\Desktop\档案汇总3\{filename}"  # 对文件执行操作，例如读取内容或处理文件
#     # 打开Excel文件
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook['经营违规']
#     # 设置页面属性
#     sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
#     sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
#     sheet.page_margins.left = 0.7
#     sheet.page_margins.right = 0.7
#     sheet.page_margins.top = 0.75
#     sheet.page_margins.bottom = 0.75
#     sheet.page_margins.header = 0.3
#     sheet.page_margins.footer = 0.3
#     print(file)

    # # 获取所有工作表
    # all_sheets = workbook.sheetnames
    #
    # # 删除除第一个工作表以外的所有工作表
    # for sheet_name in all_sheets:
    #     if sheet_name == '封面':  # 替换为你要保留的工作表的名称
    #         del workbook[sheet_name]
    # 保存修改后的Excel文件
    # workbook.save(fr"C:\Users\XYCX\Desktop\新工作簿\{filename}")

# engine = create_engine('mssql+pymssql://sa:Xy202204@LAPTOP-HHMD1A48\XYCX/证件合规信息查询?charset=utf8', echo=False)
# df = pd.read_excel(r"C:\Users\XYCX\Desktop\收集表.xlsx")
# 将日期列转换为所需格式
# df['上车时间'] = pd.to_datetime(df['上车时间'], format='%Y%m%d%H%M%S').dt.strftime('%Y-%m-%d')
# df['下车时间'] = pd.to_datetime(df['下车时间'], format='%Y%m%d%H%M%S').dt.strftime('%Y-%m-%d')
# df.to_sql("罚款明细", con=engine, index=False)

# from datetime import datetime, timedelta
#
# data_result = 1
# # 获取当前日期和时间
# current_date_time = datetime.now()
# # 将当前日期设置为当天的午夜
# midnight = current_date_time.replace(hour=0, minute=0, second=0, microsecond=0)
# # 得出前一天时间，23:59:59
# end_time = midnight - timedelta(seconds=data_result)
# # 把日期时分秒改成全是0
# start_time = end_time.replace(hour=0, minute=0, second=0, microsecond=0)
# # 格式化输出
# end_time_str = end_time.strftime('%Y-%m-%d %H:%M:%S')
# start_time_str = start_time.strftime('%Y-%m-%d %H:%M:%S')
# start_time_ms = str(int(start_time.timestamp() * 1000))
# print(type(start_time_ms))

# 爬取订单列表
# headers = {
#     'Accept': 'application/json, text/plain, */*',
#     'Accept-Language': 'zh-CN,zh;q=0.9',
#     'Connection': 'keep-alive',
#     'Content-Type': 'application/json',
#     'Cookie': '_blmsdk_as_uid=FEYImLV69RuQbHkOm69fBcO1AKpiQnURGFsFBIoSGFAkVoA3MdJe8mDUQVmFzR97nM5V7NID5k35aZbC3iPpnOueKTK3fzAvgtKTovZMEGLA68kuxMpmsLEtYHdryzAEn0dLGDMfNZkyDW56sRZJJWXp9SR2XtEiugVtsgHQee4%3D;_blmsdk_s=1929d1fd602-a26-7e4-913%7C10;_blmsdk_uid=dEHfAr%2FLil%2BZ6gbW6iPPCDPdoKXnT7JIRK1K2iKpmk1HYl5tFr3NdHcOM7tGgp5%2FbxXYyZBLtYRBvJjSTgjoNdyYTnQfC84RD24afraYPSy%2FlF7eva2AmiAeajwPtXIhRm6OBjwe2rX1AuWUSqDi6b3LTvIOoBxYKh8IKQFGD8Tju2TGM6trknyAezG8Pq8GIon7BbCO4ck9FhlMM0GpDbLj4B2u5KGvPZznxFypA1ItW3A3ld3z5KcSCu760%2FGqaJebXvPLxXYcSkCr23CwqNqZSI%2BspHjqrA60LmcrCv6ko%2Brisqcb56SZBZO%2BufrWqFq3OvGQaBAKOwwVU5O5RQ%3D%3D;_app_custom_tenant_id=800432;_assets_tenant_id=800432;_bi_tenant_id=800432;_brand_id=800432;_css_tenant_id=800432;_finance_tenant_id=800432;_order_tenant_id=800432;_tenant_id=800432;_user_expires_time=1729299396203;_yycx_admin={%22token%22:%227c5d2b348adb47f9be55df9e38001efc%22%2C%22expiration%22:1729299396203%2C%22socketTokenVo%22:{%22keyId%22:218713629%2C%22keyValue%22:%2263e88b3f78eb13c8783eda0f106a52b0%22%2C%22keyIv%22:%2293092d8fd46e0b59d32ba0c922cb7358%22%2C%22accessToken%22:%227c5d2b348adb47f9be55df9e38001efc%22}};_yycx_admin_acl=7c5d2b348adb47f9be55df9e38001efc;_yycx_admin_app_custom=7c5d2b348adb47f9be55df9e38001efc;_yycx_admin_assets=7c5d2b348adb47f9be55df9e38001efc;_yycx_admin_bi=7c5d2b348adb47f9be55df9e38001efc;_yycx_admin_ccs=7c5d2b348adb47f9be55df9e38001efc;_yycx_admin_finance=7c5d2b348adb47f9be55df9e38001efc;_yycx_admin_order=7c5d2b348adb47f9be55df9e38001efc;_yycx_socket_token_vo={%22keyId%22:218713629%2C%22keyValue%22:%2263e88b3f78eb13c8783eda0f106a52b0%22%2C%22keyIv%22:%2293092d8fd46e0b59d32ba0c922cb7358%22%2C%22accessToken%22:%227c5d2b348adb47f9be55df9e38001efc%22};acw_tc=784e2c9317292129623944392e2894b4c6716d7e88884b9cf9c79ef086c760;message_alert=0;sidebarStatus=0;stk=7c5d2b348adb47f9be55df9e38001efc;',
#     'Origin': 'https://admin.yueyuechuxing.cn',
#     'Referer': 'https://admin.yueyuechuxing.cn/order/newOrder/newList?cacheDataQuery=%7B%22pageNum%22%3A%221%2F2%22,%22key%22%3A%22params%22,%22bid%22%3A800432,%22cacheFields%22%3A%22realTimeType,pageSize,pageNum%22,%22notIncludeData%22%3A%22%7B%7D%22%7D',
#     'Sec-Fetch-Dest': 'empty',
#     'Sec-Fetch-Mode': 'cors',
#     'Sec-Fetch-Site': 'same-origin',
#     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36',
#     '_admin_current_page': '/order/newOrder/newList',
#     '_admin_eid': '800432',
#     '_admin_resource_key': 'newOrderList',
#     '_admin_session_eid': '800432',
#     '_admin_sign': '76828b6260c0588b86dff532333faab9',
#     '_admin_tk': '7c5d2b348adb47f9be55df9e38001efc',
#     '_admin_ts': str(round(time.time())),
# }
#
# from concurrent.futures import ThreadPoolExecutor, as_completed
#
# # 读取 Excel 文件
# df = pd.read_excel(r"C:\Users\XYCX\Desktop\订单号.xlsx")
# final_df = pd.DataFrame()
#
#
# # 请求的函数
# def fetch_order_detail(order_id):
#     while True:
#         try:
#             json_data = {
#                 'pageNum': 1,
#                 'showEtravelCancelOrderFlag': 0,
#                 'passCreateStartDate': 1728835200000,
#                 'passCreateEndDate': 1728912624468,
#                 'specialOrderFlag': 0,
#                 'orderId': f'{order_id}',
#             }
#             response = requests.post(
#                 'https://admin.yueyuechuxing.cn/bos/admin/v1/order/queryDetail',
#                 headers=headers,
#                 json=json_data,
#             )
#             res = response.json()['data']
#             if res is not None:
#                 return pd.DataFrame(res['items'])
#             else:
#                 print(response.json(), '失败的页数', res)
#                 time.sleep(2)
#         except Exception as c:
#             print(c)
#             time.sleep(random.uniform(3, 5))
#
#
# # 使用 ThreadPoolExecutor 进行多线程
# batch_size = 20  # 每组线程数
# for start in trange(0, len(df), batch_size):
#     with ThreadPoolExecutor(max_workers=batch_size) as executor:
#         futures = []
#         for i in range(start, min(start + batch_size, len(df))):
#             order_id = df["订单号"][i]
#             futures.append(executor.submit(fetch_order_detail, order_id))
#
#         for future in as_completed(futures):
#             new_df = future.result()
#             final_df = final_df._append(new_df, ignore_index=True)
#
#     time.sleep(random.uniform(0.1, 0.8))  # 控制请求频率
#
# # 保存最终结果到 Excel
# final_df.to_excel(r"C:\Users\XYCX\Desktop\剩余订单.xlsx", index=False)


# # 运营周报
# def just_open(filename):
#     # xlApp = Dispatch("Excel.Application")
#     xlApp = Dispatch("ket.Application")  # 用wps接口处理excel文档
#     xlApp.Visible = False  # 隐藏 Excel 窗口
#     xlBook = xlApp.Workbooks.Open(filename)
#     xlBook.Save()  # 使用 Save 方法覆盖原文件;如需要提示是否覆盖请改成SaveAs(filename)
#     xlBook.Close(False)  # 关闭文件，False不提示保存
#
#
# start_time = time.time()
# city_list = [
#     "琼桂兵团", "鲁皖兵团", "闽赣兵团", "黑吉辽兵团", "云贵兵团", "粤东兵团", "浙江兵团", "津冀兵团", "豫晋兵团", "粤西兵团",
#     "广佛兵团", "江苏兵团", "四川兵团", "湖北兵团", "陕西兵团", "深莞兵团", "中珠兵团", "湖南兵团", "上海兵团", "重庆兵团"]
# for city in city_list:
#     print(f'{emoji.emojize(":partying_face::partying_face:")}{city}{emoji.emojize(":partying_face::partying_face:")}')
#     # 加载 Excel 文件
#     file_path = r"C:\Users\XYCX\Desktop\00 【看板】罚款统计（自动报表）(2).xlsx"
#     workbook = openpyxl.load_workbook(file_path)
#     report_sheet = workbook['汇总看板']  # 选择周报表
#     # 修改 C1 和 D1 单元格的值
#     report_sheet['D4'] = city
#     new_file = rf"C:\Users\XYCX\Desktop\兵团罚款统计\{city}.xlsx"
#     workbook.save(new_file)  # 保存当前更改
#     just_open(new_file)  # 使用win32com自动打开文件并保存
#     # 以data_only=True或默认data_only=False打开会得到两种不同的结果，各自独立，即data_only=True状态下打开的，
#     # 会发现公式结果为None（空值）或者一个计算好的常数，而不会看到它原本的公式是如何。而data_only=False则只会显示公式而已。
#     # 因此，data_only=True状态下打开，如果最后用save()函数保存了，则原xlsx文件中，公式会被替换为常数结果或空值。
#     workbook = openpyxl.load_workbook(new_file, data_only=True)
#     # 删除其他工作表，只保留 "周报"
#     for sheet_name in workbook.sheetnames:
#         if sheet_name not in ['汇总看板', "罚款明细"]:
#             del workbook[sheet_name]
#     workbook.save(new_file)
# end_time = time.time()
# total_time = end_time - start_time
# print(f'{emoji.emojize(":hundred_points::hundred_points:")}耗时{total_time//60} 分钟{emoji.emojize(":hundred_points::hundred_points:")}')


# 分类移动excel表
# engine = create_engine('mssql+pymssql://sa:Xy202204@LAPTOP-HHMD1A48\XYCX/配置表?charset=utf8', echo=False)
# df = pd.read_sql("select * from dbo.兵团", con=engine)
# result = {}
# for _, row in df.iterrows():
#     team = row['兵团']
#     brand = row['品牌']
#     city = row['城市']
#     # 格式化品牌和城市的组合
#     combination = f"{brand[:2]}{city[:2]}"
#     # 如果兵团不在字典中，初始化为一个列表
#     if team not in result:
#         result[team] = []
#     # 添加格式化的名称到对应的兵团列表中
#     result[team].append(combination)
# folder = r"C:\Users\XYCX\Desktop\兵团文件夹"
# # 遍历兵团映射并移动文件
# for dic, files in result.items():各自负责的兵团城市进
#     for name in files:
#         file_name = rf"C:\Users\XYCX\Desktop\{name}.xlsx"
#         # 如果目标文件架不存在则创建
#         os.makedirs(rf"{folder}\{dic}", exist_ok=True)
#         if os.path.exists(file_name):
#             shutil.move(file_name, rf"{folder}\{dic}\{name}.xlsx")
#         else:
#             print(f"文件未找到: {file_name}")




# headers = {
#     'Accept': 'application/json, text/plain, */*',
#     'Accept-Language': 'zh-CN,zh;q=0.9',
#     'Connection': 'keep-alive',
#     'Content-Type': 'application/json',
#     'Cookie': '_blmsdk_did=0dfddaf639225d-192a242820ac8-17333273-1bcab9-192a242820ac8; _order_tenant_id=800432; _assets_tenant_id=800432; _finance_tenant_id=800432; _app_custom_tenant_id=800432; _css_tenant_id=800432; _bi_tenant_id=800432; _tenant_id=800432; stk=b41fec603a32467883db7727e242c4d7; _yycx_admin_assets=b41fec603a32467883db7727e242c4d7; _yycx_admin_finance=b41fec603a32467883db7727e242c4d7; _yycx_admin_order=b41fec603a32467883db7727e242c4d7; _yycx_admin_app_custom=b41fec603a32467883db7727e242c4d7; _yycx_admin_ccs=b41fec603a32467883db7727e242c4d7; _yycx_admin_bi=b41fec603a32467883db7727e242c4d7; _yycx_admin_acl=b41fec603a32467883db7727e242c4d7; _yycx_socket_token_vo={"keyId":2035997477,"keyValue":"ebb3d105c60c756ff2d1b3628dc4b993","keyIv":"81bd69a4b22a225c3795ef70ef701ea1","accessToken":"b41fec603a32467883db7727e242c4d7"}; _user_expires_time=1731027139465; _yycx_admin={"token":"b41fec603a32467883db7727e242c4d7","expiration":1731027139465,"socketTokenVo":{"keyId":2035997477,"keyValue":"ebb3d105c60c756ff2d1b3628dc4b993","keyIv":"81bd69a4b22a225c3795ef70ef701ea1","accessToken":"b41fec603a32467883db7727e242c4d7"}}; message_alert=0; _brand_id=800432; _blmsdk_uid=hHSNMRndp0DRgALbf6rUpWC1o4xsSVDib0hxlWl9W0YBWiwGcvXaGVu7XqvSC+tgKZ4DewQ92HL4F5Umb4dr3BiEI/mOV9LMEpvSV7MwYNeah5j/tKegrybL+qdgV9Zu+h4XPL9jAcdruqxNrs4IKyI+qdWL9TjxBu7ZRJ3b+m2qkrZM66gticNSZ/iRnm5VMf8eHzbDAWW8kgVtoRcfKLyO1veVwzNUEJ/FcCgWBabOIGq6buflT7PRQs4qz27w7jlIeXyZbFJ0HLgXS/k+zq/8+WPaNHZZLtmyV6ZOK3DmPtotvKVRCa/trtbgFBMMfbYwOJXAkhxI2rHCLie6bg==; _blmsdk_as_uid=JLqjDDO/TQlKjRfl4bEw1qvn+c4xT/CCMPtVQnHJQ695hWPMbC/xPYWEGTYRodXR/2Do9n2N2lltIpLQjzICdW8LXw5gWuq0/8uDJL/Uihs5AmyIcDdY28noqA8QgdrVzqjsW6FnG9eYtTAoeTHjOuMzqfln43mSSb9ZuMEn334=; acw_tc=2f6a1fa017309699206771374e3f917446e94f92f7df9d6ea9e5a7b729bfc9; sidebarStatus=1; _blmsdk_s=19305ebfd90-295-2f5-a95|2',
#     'Origin': 'https://admin.yueyuechuxing.cn',
#     'Referer': 'https://admin.yueyuechuxing.cn/assets/driverManage/driverList/updateDriver/300/5549026016601?tenantId=800432&adcode=440700&_fromName=driverManage&runTimeSEid=800432',
#     'Sec-Fetch-Dest': 'empty',
#     'Sec-Fetch-Mode': 'cors',
#     'Sec-Fetch-Site': 'same-origin',
#     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.0.0 Safari/537.36',
#     '_admin_current_page': '/assets/driverManage/driverList/updateDriver/300/5549026016601',
#     '_admin_eid': '800432',
#     '_admin_resource_key': 'updateDriver',
#     '_admin_session_adcode': '440700',
#     '_admin_session_eid': '800432',
#     '_admin_sign': '76828b6260c0588b86dff532333faab9',
#     '_admin_tk': 'b41fec603a32467883db7727e242c4d7',
#     '_admin_ts': '1730785072',
#     'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="101", "Google Chrome";v="101"',
#     'sec-ch-ua-mobile': '?0',
#     'sec-ch-ua-platform': '"Windows"',
# }
#
#
# df = pd.read_excel(r"C:\Users\XYCX\Desktop\江门市.xlsx")
# for i in trange(len(df)):
#     ID = str(df["司机ID"][i])
#     json_data = {
#         'driverId': ID,
#     }
#     response = requests.post(
#         'https://admin.yueyuechuxing.cn/admin/v1/assets/driver/hailing/license/info',
#         headers=headers,
#         json=json_data,
#     )
#     while True:
#         try:
#             result = response.json()["data"]
#             if result["licenseImg"]:
#                 imgurl = "https://admin.yueyuechuxing.cn" + result["licenseImg"]
#                 res = requests.post(imgurl, headers=headers).content
#                 # 将图片内容保存为文件
#                 with open(rf"C:\Users\XYCX\Desktop\推送文件\{ID}.jpg", "wb") as f:
#                     f.write(res)
#                 time.sleep(random.uniform(1, 3))
#                 break
#             else:
#                 print(f"{ID}没有照片")
#         except Exception as e:
#             print(ID)
#             traceback.print_exc()
#             time.sleep(random.uniform(1, 3))


# print(emoji.demojize("🥳"))
# print(emoji.emojize(":dove:"))

# 单条存储经营支付
# engine = create_engine('mssql+pymssql://sa:Xy202204@LAPTOP-HHMD1A48\\XYCX/证件合规信息查询?charset=utf8', echo=False)
# def writer_file(file_data, brand, city):
#     """
#     读取文件后进行数据清洗然后写入数据库
#     :return:
#     """
#     # 将时间列转换为正确的时间格式
#     file_data['上车时间'] = pd.to_datetime(file_data['上车时间'], format='%Y%m%d%H%M%S')
#     file_data['下车时间'] = pd.to_datetime(file_data['下车时间'], format='%Y%m%d%H%M%S')
#     file_data['订单完成时间'] = pd.to_datetime(file_data['订单完成时间'], format='%Y%m%d%H%M%S')
#     file_data['乘客结算时间'] = pd.to_datetime(file_data['乘客结算时间'], format='%Y%m%d%H%M%S')
#     # print(export_time)
#     file_last = file_data[["司机编号", "订单号", "机动车驾驶员姓名", "机动车驾驶证号", "车辆号牌", "上车地点", "上车时间",
#                            "下车地点", "下车时间", "订单完成时间", "乘客结算时间", "实收金额", "应收金额"]].copy()
#     file_last.columns = ["司机ID", "订单号", "姓名", "身份证号", "车牌号", "上车地点", "上车时间", "下车地点", "下车时间",
#                          "订单完成时间", "乘客结算时间", "实收金额", "应收金额"]
#     file_last.insert(0, '品牌', brand)
#     file_last.insert(1, '城市', city)
#     file_last['导出时间'] = "2024-12-10"
#     file_last['传输时间'] = "2024-01-29"
#     file_last.to_sql('经营支付', con=engine, if_exists='append', index=False)
#     print("存完了")
# datn = pd.read_excel(r"C:\Users\XYCX\Downloads\经营支付.xlsx",
#                      dtype={'司机编号': str, '订单号': str, '机动车驾驶证号': str})
# writer_file(datn, '喜行约车', "温州市")


# # 起始时间
# start_date = datetime.datetime.now() - datetime.timedelta(days=1)
# start_Time = start_date.strftime('%Y-%m-%d')
# # 起始时间戳
# start_time_obj = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
# startTime = int(start_time_obj.timestamp() * 1000)
#
# # 截止时间
# end_date = datetime.datetime.now() - datetime.timedelta(days=1)
# end_Time = end_date.strftime('%Y-%m-%d')
# # 截止时间戳
# end_time_obj = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
# endTime = int(end_time_obj.timestamp() * 1000)

# 输出结果
# print("起始时间", start_Time)
# print("起始时间戳", startTime)
# print("截止时间", end_Time)
# print("截止时间戳", endTime)


# data_result = 1
# # 起始日期
# today = (datetime.datetime.now() - datetime.timedelta(days=data_result)).strftime('%Y-%m-%d')
# # 今天的时间戳
# endTime = int(time.mktime(time.strptime("%s 23:59:59" % today, "%Y-%m-%d %H:%M:%S"))) * 1000
# # 起始时间
# this_month_start = (datetime.datetime.now() - datetime.timedelta(days=data_result)).strftime('%Y-%m-%d')
# # this_month_start的时间戳
# startTime = int(time.mktime(time.strptime("%s 00:00:00" % this_month_start, "%Y-%m-%d %H:%M:%S"))) * 1000
# print(today)
# print(endTime)
# print(this_month_start)
# print(startTime)

# 钉钉发消息
# from dingtalkchatbot.chatbot import DingtalkChatbot, ActionCard, CardItem
# def dingtalk_robot(webhook, secret):
#     dogBOSS = DingtalkChatbot(webhook, secret)
#     red_msg = '<font color="#dd0000">级别:危险</font>'
#     orange_msg = '<font color="#FFA500">级别:警告</font>'
#
#     now_time = datetime.now().strftime('%Y.%m.%d %H:%M:%S')
#     url = 'https://blog.csdn.net/qq_46158060?type=blog'
#     dogBOSS.send_markdown(
#         title=f'来自机器人的提醒',
#         text=f'### **我是主内容的第一行**\n'
#              f'**{red_msg}**\n\n'
#              f'**{orange_msg}**\n\n'
#              f'**发送时间:**  {now_time}\n\n', is_at_all=True)  # 是否@所有人
# webhook = 'https://oapi.dingtalk.com/robot/send?access_token=9c05b114ae403d03d257881811939035f2795281891a44d493d2341924ce401c'
# secrets = 'SEC2d87ab004fa71c31cf402f8800108625a68e5a678e82fe9e5db1771b83b08f07'
# dingtalk_robot(webhook=webhook, secret=secrets)


# start_date = '2024-11-08'
# end_date = '2024-11-09'
# # 使用 strptime 将字符串转换为 datetime 对象
# start_Time = datetime.strptime(start_date, '%Y-%m-%d')
# end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
# while True:
#     if start_Time.month == 1 and start_Time.day == 1:
#         break
#
#     if end_datetime.day == 1:
#         start_Time = end_datetime.replace(hour=0, minute=0, second=0)
#         end_Time = end_datetime.replace(hour=23, minute=59, second=59)
#     else:
#         end_Time = end_datetime.replace(hour=23, minute=59, second=59)
#         start_Time -= timedelta(days=2)
#
#     print(start_Time.strftime('%Y-%m-%d %H:%M:%S'), end_Time.strftime('%Y-%m-%d %H:%M:%S'))
#     time.sleep(0.3)
#
#     # 更新下一轮循环的起始时间和结束时间
#     if end_datetime.day == 1:
#         end_datetime -= timedelta(days=1)
#         start_Time = end_datetime.replace(hour=0, minute=0, second=0)
#     else:
#         end_datetime -= timedelta(days=2)
# print(emoji.demojize("👻"))


# headers = {
#     'accept': 'application/json, text/plain, */*',
#     'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
#     'content-type': 'application/json;charset=UTF-8',
#     'origin': 'https://touch.qunar.com',
#     'priority': 'u=1, i',
#     'referer': 'https://touch.qunar.com/hoteldetail?seq=beijing_city_12038&checkInDate=2025-03-29&checkOutDate=2025-03-30&bd_source=_600010157&notLogin=true',
#     'sec-ch-ua': '"Chromium";v="134", "Not:A-Brand";v="24", "Microsoft Edge";v="134"',
#     'sec-ch-ua-mobile': '?0',
#     'sec-ch-ua-platform': '"Windows"',
#     'sec-fetch-dest': 'empty',
#     'sec-fetch-mode': 'cors',
#     'sec-fetch-site': 'same-origin',
#     'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36 Edg/134.0.0.0',
#     'cookie': '__qt=v1%7CVTJGc2RHVmtYMTl2OEJsb3J6RGZrTkdVK0FhRERuSVhpdmdmY0piT1QyeWoxUU9TUEd6d1ZoanlsNUh2ZTFvZk95SWl6Q1hhaTVvT1B6d0taOUErYm5uSUo0c3hTRDhYc3lyL1Zra3FVNGt2UlhzK1FDV1Z1OUN2aU5mQW9uREl6ZThtcnA5YUFuTzhwdE5tc2tvN3EvR1RIWUlSanJ2TWxEc29yeG5uMnA5VzVnaUY5Y2VNSlpVZkQ5b0E4eFEzRExtWE54VWpLOXVtaVJFbzdhMGhyclhGTGYrK3JOYVg1Q0h6WWlTSmFkOD0%3D%7C1743210799622%7CVTJGc2RHVmtYMS9DcHBTaG9vaTNiaUF4MWx4QVh5L1FzRGtCZHdsMmRhVFlDWThrTmNXc09rNUlac1cyVnJQYnJsM3NsaElEVXU3SHVmdzNhWHdRalE9PQ%3D%3D%7CVTJGc2RHVmtYMTg3d2pEWlVBbml1ZmlQV3lZVVBZNnlSUXRlTXdYU2xienpiOGo5VUxPTTM0YnlOaXdGOVlyemYrR1ROMytyVzVMeTJVTUU1S01SRnoxK0NJcVVLSjB1aGRmeHJ6YWsrY0preGtzbGFqUmNheTZKSWV5d2lrdUswc2R6MmtLUmpPU3cyMEYrWjgwMWVjS24xUlFLSTNGSks4NU4ydmx0Tkt5V2p6K2U5TUlDR3FsMWN3Zk5IeWQ0dUs1eW1YcUlTcnhXajZLOHZiY0V6UWcxZWsxNFYzMWNwdW9scmlLeUc2blBiQUpZK0dKM3ZUeHZYTU9LMlMxYjl6WG9JbFg2VHI2anNTVmJOK2hReGVyWk83QVIxMmNPV2c2MGJYMmcwM2xmV2dYM3ZheUp5WGU3Y3RVZjFyM0xienVUR25KUmsxeFljQUZIOG45OUlTcFFBRTdEZGRiTVRHb0puV3JBczM3a294MTdBSFRIRkcxdXlucGZ0aVFKRktBeG9ybmJvWFdvT2F2cWRacFpkVXE5VmJBaVdweVlueDBQT3FDNlVXbXdnYVYxTlZPSlBqK3VHSWlDZk5VVjJrSGY1UmR5K01kNEpKS1RBSkRSNHczQlJNS3hWNnN5Nmw3MzdTSi80bzFPSDRTNTVHMFlXUzd4SGhNQlE2cE13T1ZuVUwrV0R5SmoyalMzWFN0YkROamNxY0FxZ21OTStxT3BxQlpFTDI4dVZjWStvcHh2aisrT2U0aTVVYnpIZWYrbDFpa1c1UGowNS9PazFKQ0FGNEFtTnNYcExkclFqT01QaDNIUWNtSXcxYVpkK21HK1oyeWRYcC9idHhMQjdrcTVGcjdaVGlBZy9IN0xrUmJhWU5zNlFqVXJpVlZZRVZLM1gvMW9hZGM5MllOWHhtekpUdzNtMWtQNDBKWEQ%3D; QN1=0001718030686cee7550a567; QN99=1299; qunar-assist={%22version%22:%2220211215173359.925%22%2C%22show%22:false%2C%22audio%22:false%2C%22speed%22:%22middle%22%2C%22zomm%22:1%2C%22cursor%22:false%2C%22pointer%22:false%2C%22bigtext%22:false%2C%22overead%22:false%2C%22readscreen%22:false%2C%22theme%22:%22default%22}; QN205=s%3Dbing; QN277=s%3Dbing; csrfToken=toJfWM9whaeKuZQWXVmS6tWAxxcalssY; _i=VInJOQ3N1RJqhmnxMmK2rHm4AsJq; _vi=CwccYHPBT0Qkw8Wx0KBihUVUcDMAPC8FLBXsNoa7VkYLSDLxz2BVknMtz7TRZ94_kFhQDC2oM7gL9N2SJSFXsuAF3PiRTBHdkgLfC1cNbe-esrTPMh-fgFADorPAJNbnwWBZDxr292ytVa4A0Kq13oHTYyy8GIItE3qmZTwGLoP2; QN601=e80b55971e6efacfbed444959e8b7708; QN269=B15947200C3A11F095F586C38D9A27DE; QN48=00010a802f106cee755811fa; quinn=296caca338c3f3fffc20ae85677274facada1a0f3056fd5aacad79f3b69b5fe6fbe72ee6a2d0649834fe42bc6671c3e4; fid=6746e437-5869-41a0-b681-ce88fccb0005; QN271=d2ebe6a8-7e44-40c0-b843-bd6b4454a2be; HN1=v18b014ec6a90afc2c560373ccf8916adb; HN2=qkrqunsglcgcu; ctt_june=1683616182042##iK3wWKaAauPwawPwasa%2BXstnXsvNVKPnX2jsWKXAXKkTaRt%3Da2EGX%3D3mWsTTiK3siK3saKg%3Das3naRj%2BaRPsWUPwaUvt; QN25=1b7b7fc7-7811-4255-a777-97506dc10a1e-9f992f90; RT=s=1743210680591&r=https%3A%2F%2Fflight.qunar.com%2F; QN66=_600010157; QN300=_600010157; QN271AC=register_pc; QN271SL=b096601aa097563c4618853c43a81022; QN271RC=b096601aa097563c4618853c43a81022; _s=s_GDYXVTF26PPUUZLELJNACGPTDQ; _v=vR8n_NLJ3nggJ_ypoXGNi5ejl9839eHN1AOz5uFENjP8c-s2DhSbPGdnMsndZVUnWxDldB81Nnj6D3QoW0idPLULfwHS6HoEMTZA1CDvUax-h12gT3eBf9yfsGsTUZT_pt6H6t5JIZoRZZxkhrYACmyyYL9JD9x69FBfyDUACqs3; _t=29292552; _q=U.wipsuez4482; QN42=%E5%8E%BB%E5%93%AA%E5%84%BF%E7%94%A8%E6%88%B7; ctf_june=1683616182042##iK3wWRXnawPwawPwasfhaDawERXAESEIEDa%3DERWDEKiRXKt%3DaSoRW2ETXs3NiK3siK3saKg%3Das3naRgmWKXOVuPwaUvt; QN267=1575743149015b90ef; cs_june=189332180a402d4fb1dd870ff683c170ea81caf6da631853c08f7601cf324a41adb58e6da00d7d905a94fb306655d37838226ecec88d0281757edfaefe0a0076b17c80df7eee7c02a9c1a6a5b97c1179418c8c2d2cfd3ac66288362bd2ded47f5a737ae180251ef5be23400b098dd8ca',
# }
#
# json_data = {
#     'seq': 'beijing_city_12038',
#     'page': 1,
#     'enBella': '1683616182042##879a71ed2bde797bf830c122360f2b9910e4d55a##iKohiK3wgMkMf-i0gUPwaUPsXuPwa5ksf-3bg-kbj-3bjOFeiKiIiK3wiKiRiK3wgI0SjOFLcwPwaUPsXuPwaStsWR20aSa0aSawE238XsGIXsP+WsGDWRGhiK3siK3saKg=as3naRt8WRj+ahPwaUPwXwPwa5Wpy-iLf-20aS30a=D0aSiUW9jnWRT0WSWHVR2=fRWUjKa8WIjAjsg+VRg8WSgnahPwaUPwXwPwa5EQoIn0iK3wiKWTiK3w-wkGWuPmXwPNWwkGWhkhXukTXwkGawPmahPNahkGWuPmEukhXUkGWuPNawkTXukGWuPmWhkhEUkGVuPmWuPNaUkGWukhXuPNWwkGawPmahPNauPwaUPwXwPwaMe0d-oxgMEsiK3wiKWTiK3wiPPNiPDOiKtOJhkGWuPmXwPNWwkGWhkhXukTXwkGWuPmEhPNWwkGWwkTXUPNVukGWukhXuPNWw20EKX0X2X0VDX0EKg0X2X0XPP0aS30a2a0aSi2f-WSgM08oI0xcUPwaUPsXuPwaUkGWuPmXwPNWwkGWhkhXukTXwkGWukhEhkTVukGVhPNVukhVukGWukTEukhEukGVukTWUPmWUt0EKP0VDa0VKg0EKX0X2D0XPa0EKP0VDX0VKg0EKg0XP30VK20EKP0X2D0VKgQiPPAiPGGiK2miPP+iPiDiK2niPPmiKtniK2=iPPAiK28iKtmiPPNiPDwiKt=iPPmiPGGiPDwiPkIiPiRiKHRiK3wiKiRiK3wyIFsohPwaUPsXuPwa5Exo9Wpq5GAcMGwqMWxcuPwaUPwXwPwa5WSgM08oGWwjwPwaUPsXuPAXUPwa5GHc5Xbg-kbj-3bjOFeiKiIgwPwaUPwXwPwa5GQc9osq5GAcMGwd5pbjO10aS30WPX0W=Xt##vbaKQyqUpdZYmAKndSQjr##seq,page',
#     'Bella': {
#         'b': {
#             'seq': 'beijing_city_12038',
#             'page': 1,
#         },
#     },
# }
#
# response = requests.post('https://touch.qunar.com/hotelcn/api/commentlist', headers=headers, json=json_data)
# print(response.json())
# headers = {
#     'accept': 'application/json, text/plain, */*',
#     'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
#     'content-type': 'application/json;charset=UTF-8',
#     'priority': 'u=1, i',
#     'sec-ch-ua': '"Chromium";v="134", "Not:A-Brand";v="24", "Microsoft Edge";v="134"',
#     'sec-ch-ua-mobile': '?0',
#     'sec-ch-ua-platform': '"Windows"',
#     'sec-fetch-dest': 'empty',
#     'sec-fetch-mode': 'cors',
#     'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36 Edg/134.0.0.0',
# }
# data = {"key": "b0a142af4f40b355efa3a51d037a9ad8", "tag": "", "message": "", "message.disposeId": "",
#         "message.gdViolationId": "", "message.plateNo": "", "message.disposeType": "", "message.disposeStatus": "",
#         "message.disposeDesc": "", "message.disposeStartTime": "", "message.disposeEndTime": "",
#         "message.punishAmount": "", "message.amapOrderId": "", "message.unionId": "", "message.channel": "",
#         "message.gmtModified": ""}
# res = requests.get("https://www.runoob.com/try/ajax/json_demo.json", headers=headers)
# res = requests.post("https://saas.amap.com/ws/service-control/sync/blm/data", headers=headers, json=data)
# print(res.json()["sites"][0]["info"][0])

#powerbi获取token
# 配置信息
# PBI_CLIENT_ID = "fc9562ac-ea68-4da9-872b-d5500b2c54a7"  # 替换为你的客户端ID
# PBI_CLIENT_SECRET = "2qtpu-xGeQQ2~hy6R9qrUq~85~~oN4-t4r"  # 替换为你的客户端密钥
# PBI_TENANT_ID = "a57dcbde-032f-4047-a9bc-872800adb61a"  # 替换为你的租户ID
# PBI_REPORT_ID = "8ed35810-ed4c-407a-a118-19c500d03756"  # 替换为你的报表ID
# PBI_GROUP_ID = "01c5b293-3cfc-47cb-837a-07081108889d"  # 替换为你的工作区ID
# def get_embed_token():
#     # 获取Azure AD令牌
#     auth_url = f"https://login.chinacloudapi.cn/{PBI_TENANT_ID}/oauth2/token"
#     auth_data = {
#         "grant_type": "client_credentials",
#         "resource": "https://analysis.chinacloudapi.cn/powerbi/api",
#         "client_id": PBI_CLIENT_ID,
#         "client_secret": PBI_CLIENT_SECRET
#     }
#     headers = {
#         "Content-Type": "application/x-www-form-urlencoded"
#     }
#     auth_response = requests.post(auth_url, headers=headers, data=auth_data)
#     print(auth_response.json())
#     access_token = auth_response.json().get("access_token")
#     headers = {
#         "Authorization": f"Bearer {access_token}"
#     }
#     # 2. 获取嵌入令牌 groups后面是工作区id
#     url = f"https://api.powerbi.cn/v1.0/myorg/groups/{PBI_GROUP_ID}/reports"
#     res = requests.get(url, headers=headers)
#     reports = res.json()
#     print(reports)
#     # 获取Embedtoken
#     new_url = f"https://api.powerbi.cn/v1.0/myorg/groups/{PBI_GROUP_ID}/reports/{PBI_REPORT_ID}/GenerateToken"
#     headers = {
#         "Content-Type": "application/json",
#         "Authorization": f"Bearer {access_token}",
#         "Accept": "application/json"
#     }
#     data = {"accessLevel": "View", "allowSaveAs": "true"}
#     response = requests.post(new_url, headers=headers, json=data)
#     print(response.json())
# get_embed_token()


engine = create_engine('mysql+pymysql://root:M25xx%ycl@HgDataBase:4408/sophon?charset=utf8mb4', echo=False)
# # df = pd.read_sql("select * from `sophon`.accounts", con=engine)
# # print(df)
# dit = {
#     '喜行约车': {'file': 'Profile 1', 'user': 'admin', 'eid': '800432'},
#     '神州专车': {'file': 'Profile 2', 'user': 'shenzhou', 'eid': '800447'},
#     '江南出行': {'file': 'Profile 3', 'user': 'jiangnan', 'eid': '800188'},
#     '星徽出行': {'file': 'Profile 4', 'user': 'xinghui', 'eid': '800161'},
#     '蛋卷出行': {'file': 'Profile 5', 'user': 'danjuan', 'eid': '800761'}
# }
# def get_cookie():
#     while True:
#         ck_df = pd.DataFrame()
#         for i in dit:
#             cookie_path = browser_cookie3.chrome(
#                 cookie_file=rf'C:\Users\XYCX\AppData\Local\Google\Chrome\User Data\{dit[i]["file"]}\Network\Cookies',
#                 domain_name='admin.yueyuechuxing.cn')
#             string, admin_tk = '', 'error'
#             for cookie_file in cookie_path:
#                 if cookie_file.name == 'stk':
#                     string = string + cookie_file.name + '=' + cookie_file.value + ';'
#                     admin_tk = cookie_file.value
#                 else:
#                     string = string + cookie_file.name + '=' + cookie_file.value + ';'
#             ck_df = ck_df._append(
#                 pd.DataFrame({'account': '小周', 'brand': i, 'eid': dit[i]["eid"], 'session_eid': dit[i]["eid"],
#                               'token': admin_tk, 'cookie': string}, index=pd.Index([len(ck_df)])))
#             time.sleep(random.uniform(0.8, 1))
#         if 'error' in ck_df['token'].values:
#             print(f'cookie错误，正在重新写入{emoji.emojize(":warning::warning::warning:")}')
#             time.sleep(random.uniform(2, 3))
#         else:
#             ck_df.to_sql('accounts', con=engine, if_exists='append', index=False)  # append 累加replace重新创表
#             print(f'已写入数据库{emoji.emojize(":partying_face::partying_face::partying_face:")}')
#             break

# def get_cookie():
#     while True:
#         ck_df = pd.DataFrame()
#         for i in dit:
#             cookie_path = browser_cookie3.chrome(
#                 cookie_file=rf'C:\Users\XYCX\AppData\Local\Google\Chrome\User Data\{dit[i]["file"]}\Network\Cookies',
#                 domain_name='admin.yueyuechuxing.cn')
#             string, admin_tk = '', 'error'
#             for cookie_file in cookie_path:
#                 if cookie_file.name == 'stk':
#                     string = string + cookie_file.name + '=' + cookie_file.value + ';'
#                     admin_tk = cookie_file.value
#                 else:
#                     string = string + cookie_file.name + '=' + cookie_file.value + ';'
#             ck_df = ck_df._append(
#                 pd.DataFrame({'Account': dit[i]["user"], '_admin_tk': admin_tk, 'cookie': string, 'eid': dit[i]["eid"],
#                               'sesseid': dit[i]["sesseid"], 'brand': i},
#                              index=pd.Index([len(ck_df)])))
#             time.sleep(random.uniform(0.8, 1))
#         if 'error' in ck_df['_admin_tk'].values:
#             print(f'cookie错误，正在重新写入{emoji.emojize(":warning::warning::warning:")}')
#             time.sleep(random.uniform(2, 3))
#         else:
#             ck_df.to_sql('小周账号信息配置表', con=engine, if_exists='append', index=False)  # append 累加replace重新创表
#             print(f'已写入数据库{emoji.emojize(":partying_face::partying_face::partying_face:")}')
#             break
#
# get_cookie()

# 新增城市powerbi网站账号
# from werkzeug.security import generate_password_hash
# # 输入用户名和密码
# username = "hegui"
# password = "Xy123456"
# # 生成哈希
# hash_value = generate_password_hash(password)
# print(f"用户名: {username}")
# print(f"密码哈希: {hash_value}")

# column_bytes = b'\xe6\x9c\x80\xe5\x90\x8e\xe4\xb8\x80\xe6\xac\xa1\xe5\x87\xba\xe8\xbd\xa6\xe6\x97\xb6\xe9\x97\xb4'
# column_name = column_bytes.decode('utf-8')  # 解码为字符串
# print(column_name)
# from sqlalchemy import create_engine, text

# import base64
# import json
# def base64_api(i):
#     base64_data = base64.b64encode(i)
#     b64 = base64_data.decode()
#     data = {"username": 'Xy123', "password": 'Xy123456', "typeid": 3, "image": b64}
#     result = json.loads(requests.post("http://api.ttshitu.com/predict", json=data).text)
#     if result['success']:
#         return result["data"]["result"]
#     else:
#         return result["message"]
#     return ""
#
#
# from base64 import b64encode
# from Crypto.PublicKey import RSA
# from Crypto.Cipher import PKCS1_v1_5
# # 配置公钥 (Base64格式)
# public_key_base64 = "MFwwDQYJKoZIhvcNAQEBBQADSwAwSAJBAKWGr8gd+hKMBQ+LQLsy0QDI0u5sK2A9ZrUt1F6yOW5TYvl9Wds498DKXieE0SjHKe1Ygoee3gnmjua8Cy7utS8CAwEAAQ=="
#
# # 将Base64公钥转换为DER格式
# public_key_der = bytes.fromhex('305c300d06092a864886f70d0101010500034b003048024100a586afc81dfa128c050f8b40bb32d100c8d2ee6c2b603d66b52dd45eb2396e5362f97d59db38f7c0ca5e2784d128c729ed5882879ede09e68ee6bc0b2eeeb52f0203010001')
# # 注意：上面的十六进制字符串是通过解码Base64公钥得到的固定值
# # 如果更换公钥，需要重新生成这个DER格式的二进制数据
#
# # 创建RSA公钥对象
# key = RSA.import_key(public_key_der)
# cipher = PKCS1_v1_5.new(key)
#
# # RSA加密函数
# def rsa_encrypt(plaintext):
#     encrypted = cipher.encrypt(plaintext.encode())
#     return b64encode(encrypted).decode()
#
# headers = {
#     'Accept': '*/*',
#     'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
#     'Authorization': 'null',
#     'Connection': 'keep-alive',
#     'Origin': 'https://60.190.56.95',
#     'Referer': 'https://60.190.56.95/',
#     'Sec-Fetch-Dest': 'empty',
#     'Sec-Fetch-Mode': 'cors',
#     'Sec-Fetch-Site': 'same-origin',
#     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36 Edg/136.0.0.0',
#     'content-type': 'application/json',
#     'sec-ch-ua': '"Chromium";v="136", "Microsoft Edge";v="136", "Not.A/Brand";v="99"',
#     'sec-ch-ua-mobile': '?0',
#     'sec-ch-ua-platform': '"Windows"',
# }
#
# response = requests.get(
#     f'https://60.190.56.95/sunlandapi/sso/api/v1/captcha/login/verifyCode?t={round(time.time()*1000)}',
#     headers=headers,
# )
# captcha = response.headers.get("set-cookie").split(';')[0]
# pic = response.content
# code = base64_api(pic)
# print(code)
# headers["cookie"] = captcha
# json_data = {
#     'username': rsa_encrypt('xixing'),
#     'password': rsa_encrypt('Xixing^%$792'),
#     'captcha': code,
#     'notecode': '123456',
# }
# print(headers)
# print(json_data)
# res = requests.post('https://60.190.56.95/api/session', headers=headers, json=json_data)
# print(res.status_code)
# print(res.headers.get("set-cookie"))

# 国交新增数据
engine = create_engine('mysql+pymysql://root:M25xx%ycl@HgDataBase:4408/sophon?charset=utf8mb4', echo=False)
with engine.connect() as connection:
    info = connection.execute(text("SELECT field_cn, field_en FROM `edsac`.`fields_info` WHERE table_name='tiaffic_month'"))
    field_dict = {row[0]: row[1] for row in info.fetchall()}
df = pd.read_excel(r"C:\Users\XYCX\Desktop\2025年5月喜行约车各城市网约车合规情况(1).xlsx")
df = df.rename(columns=field_dict)
print(field_dict)
df.to_sql('tiaffic_month', con=engine, if_exists='append', index=False)
