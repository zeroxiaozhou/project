import time
import traceback

import emoji
import random
import hashlib
import os
import shutil
import requests
from datetime import datetime
from openpyxl import load_workbook
import pandas as pd
from sqlalchemy import create_engine

# engine = create_engine('mssql+pymssql://sa:Xy202204@LAPTOP-HHMD1A48\XYCX/证件合规信息查询?charset=utf8', echo=False)
engine = create_engine('mysql+pymysql://root:M25xx%ycl@HgDataBase:4408/sophon', echo=False)
# 获取当前日期
today = datetime.now()
# 格式化日期为MM-DD
formatted_date = today.strftime('%m-%d')
lst = ["喜行厦门", "喜行江门", "喜行长沙", "喜行揭阳", "喜行大庆", "喜行嘉兴", "喜行中山", "喜行肇庆",
       "喜行洛阳", "喜行重庆", "喜行烟台", "喜行西安", "喜行梅州", "喜行六安", "喜行石家庄", "喜行常州", "喜行西安", "喜行肇庆",
       "喜行南通", "喜行沈阳", "喜行成都", "喜行福州", "喜行桂林", "喜行贵阳", "喜行长春", "喜行天津"]

def optafile():
    df = pd.read_excel(rf"C:\Users\XYCX\Desktop\总推送数据{formatted_date}.xlsx")
    # 按品牌和城市分组
    grouped = df.groupby(['品牌', '城市'])
    # 保存每个分组到一个新的文件
    for name_lst, group in grouped:
        brand, city = name_lst
        # if any(city_name in city for city_name in lst):
        #     filename = fr'C:\Users\XYCX\Desktop\待推送文件\{brand[0:2]}{city[0:-1]}新增推送{formatted_date}.xlsx'
        #     group.to_excel(filename, index=False)
        # else:
        filename = fr'C:\Users\XYCX\Desktop\推送文件\{brand[0:2]}{city[0:-1]}新增推送{formatted_date}.xlsx'
        group.to_excel(filename, index=False)

    for filename in os.listdir(r'C:\Users\XYCX\Desktop\推送文件'):
        wb = load_workbook(rf'C:\Users\XYCX\Desktop\推送文件\{filename}')
        original_ws = wb.active
        # 创建一个新的工作表
        new_ws = wb.create_sheet(title='Sheet2')
        # 把新工作表移动到工作簿的最前面
        wb.move_sheet(new_ws, offset=-len(wb.worksheets) + 1)
        # 写入固定内容
        new_ws['A1'] = """导入说明：
1.模版字段信息（表头）不可增加或删除。前3行不可删除，需要导入的数据从第4行开始。
2.红色字段为必填信息；蓝色字段为示例数据，不可删除或修改。"""
        new_ws['A2'] = '司机编号'
        new_ws['A3'] = '28178517147'
        # 复制司机ID到新的工作表
        for row in original_ws.iter_rows(min_row=2, max_row=original_ws.max_row, min_col=1, max_col=1):
            for cell in row:
                new_ws.append([str(cell.value)])
        # 保存修改后的工作簿
        wb.save(rf'C:\Users\XYCX\Desktop\推送文件\{filename}')
    print("=============分类完成=============")


def calculate_md5(data: bytes) -> str:
    md5_hash = hashlib.md5()
    md5_hash.update(data)
    return md5_hash.hexdigest()


def detail(file_path, header):
    """
    :param :
    :return: 返回的是文件的url
    """
    while True:
        with open(file_path, 'rb') as file:
            file_data = file.read()
        file = {'file': (file_path.rsplit('\\', 1)[1], file_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {
            'md5': calculate_md5(file_data),
            'name': file_path.rsplit('\\', 1)[1],
            'contentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'group': 'private',
            'module': 'file',
            'fileLength': os.path.getsize(file_path),
            'mode': 'VPC_READ'
        }
        response = requests.post(
            'https://admin.yueyuechuxing.cn/admin/v1/common/upload/file',
            headers=header,
            data=data,
            files=file
        )
        time.sleep(random.uniform(1.5, 2))
        try:
            if response.json()["code"] == 1:
                print(f'上传完成{emoji.emojize(":incoming_envelope::incoming_envelope:")}')
                pdf_url = response.json()["data"]["url"]
                return pdf_url
            else:
                print(f'上传失败{emoji.emojize(":envelope::envelope:")}', response.json())
                time.sleep(random.uniform(1.5, 2))
        except Exception as e:
            print(f'上传文件报错原因{emoji.emojize(":warning::warning:")}', e)
            time.sleep(random.uniform(1, 1.5))


def parse(file, header, city_name):
    """
    上传文件后需要点击下一步进行上传，然后再点击完善
    """
    file_url = detail(file, header)
    json_data = {
        'adcode': city_dic[city_name],
        'pushStatus': '1',
        'selectType': '0',
        'selectDataType': '1',
        'fileUrl': file_url,
        'relationFlag': 1,
        'agreementId': 10,
        'agreementVersion': 1679563874669,
        'agreementFlag': True,
    }
    response = requests.post('https://admin.yueyuechuxing.cn/bos/admin/v2/sv/push/newCreateTask',
                             headers=header, json=json_data)
    # 返回数据中获取到任务ID：taskId
    result = response.json()
    # print("上传文件后点击了下一步", )
    time.sleep(random.uniform(1, 1.5))
    # 上传完文件又点击了下一步之后,判断有没有传输完成，传输完成后才可以继续点击下一步
    while True:
        res = requests.post('https://admin.yueyuechuxing.cn/admin/v1/opsx/batchimport/progressBar', headers=header,
                            json={"taskId": result["data"]["importTaskId"]})
        res_data = res.json()
        # print(res_data)
        if res_data["data"]["taskStatus"] == 1:
            print(f'传输完成{emoji.emojize(":smiling_face_with_hearts::smiling_face_with_hearts:")}')
            break
        else:
            print(f'{emoji.emojize(":sweat_droplets::sweat_droplets:")}还在传输中'
                  f'{emoji.emojize(":anger_symbol::anger_symbol:")}, {res_data}')
            time.sleep(random.uniform(10, 13))
    # 传输完成后点击下一步
    requests.post('https://admin.yueyuechuxing.cn/bos/admin/v2/sv/push/asyncCheckResult', headers=header,
                  json={"taskId": result["data"]["taskId"]})
    # print("上传文件点完下一步后的下一步", q.json())
    time.sleep(random.uniform(10, 13))
    while True:
        # 点击刷新查看是否校验完成
        new_res = requests.post('https://admin.yueyuechuxing.cn/bos/admin/v2/sv/push/findById',
                                headers=header, json={"taskId": result["data"]["taskId"]})
        new_data = new_res.json()
        try:
            if new_data["data"]["taskStatus"] == 3:
                print(f'校验完毕{emoji.emojize(":smiling_face_with_hearts::smiling_face_with_hearts:")}')
                break
            else:
                # print(f'{emoji.emojize(":sweat_droplets::sweat_droplets::sweat_droplets:")}正在校验'
                #       f'{emoji.emojize(":anger_symbol::anger_symbol::anger_symbol::anger_symbol:")}, {new_data}')
                time.sleep(random.uniform(10, 20))
        except Exception as e:
            print('new_data报错原因', e, "报错的数据", new_data)
            time.sleep(random.uniform(10, 20))
    # 再点击完善
    up_res = requests.post('https://admin.yueyuechuxing.cn/bos/admin/v2/sv/push/asyncTestGroupResult', headers=header,
                           json={"testGroupFlag": 0, "taskId": result["data"]["taskId"],
                                 "adcode": city_dic[city_name], "dataTagsMap": {}})
    print(f'任务ID：{result["data"]["taskId"]},', up_res.json())



dit = {
    '蛋卷': {'file': 'Profile 5', 'user': '蛋卷出行', 'eid': '800761'},
    '喜行': {'file': 'Profile 1', 'user': '喜行约车', 'eid': '800432'},
    '神州': {'file': 'Profile 2', 'user': '神州专车', 'eid': '800447'},
    '星徽': {'file': 'Profile 4', 'user': '星徽出行', 'eid': '800161'},
    '江南': {'file': 'Profile 3', 'user': '江南出行', 'eid': '800188'},
}
sql = "select brand AS 品牌,city AS 城市,adcode AS 行政区划代码 from `edsac`.push_city_info"
da = pd.read_sql(sql, con=engine)
da['品牌城市'] = da['品牌'].str[:2] + da['城市'].str[:-1]
city_dic = da.set_index('品牌城市')['行政区划代码'].to_dict()
# 先把推送文件分类好
optafile()
# quit()

name_list = []
for j in os.listdir(r"C:\Users\XYCX\Desktop\推送文件"):
    query = "SELECT * FROM accounts WHERE account='小周' AND brand LIKE %s"
    pattern = f"%{dit[j[0:2]]['user']}%"  # 例如："%神州专车%"
    ck = pd.read_sql(query, con=engine, params=(pattern,))
    cookie = ck['cookie'][0]
    _admin_tk = ck['token'][0]
    name = j[:-14]
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36 Edg/115.0.1901.183',
        'cookie': cookie,
        '_admin_eid': dit[j[0:2]]["eid"],
        '_admin_session_eid': dit[j[0:2]]["eid"],
        '_admin_sign': '76828b6260c0588b86dff532333faab9',
        '_admin_tk': _admin_tk,
        '_admin_ts': str(round(time.time())),
    }
    if any(city in name for city in lst):
        print(f"{name}跳过")
        shutil.move(rf"C:\Users\XYCX\Desktop\推送文件\{j}", rf"C:\Users\XYCX\Desktop\待推送文件\{j}")
        continue
    try:
        a = city_dic[name]
        print(f'========{emoji.emojize(":partying_face:")}{name}开始{emoji.emojize(":partying_face:")}============')
        print(rf"C:\Users\XYCX\Desktop\推送文件\{j}")
        parse(rf"C:\Users\XYCX\Desktop\推送文件\{j}", headers, name)
        shutil.move(rf"C:\Users\XYCX\Desktop\推送文件\{j}", rf"D:\数据\新增推送\{j}")
    except KeyError:
        # traceback.print_exc()
        name_list.append(name)
        continue
print(f'{emoji.emojize(":anger_symbol::anger_symbol:")}这些城市没有：{name_list}')
